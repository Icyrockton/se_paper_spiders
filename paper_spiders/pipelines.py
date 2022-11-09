# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import scrapy
from itemadapter import ItemAdapter
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Dict, Tuple, List
from openpyxl.styles import DEFAULT_FONT, Font, NamedStyle
from openpyxl.cell import Cell
from .utils.paperlist import paper_list


def config_style():
    style = NamedStyle(name='paper_style')
    style.font = Font(name='等线', sz=14)
    return style


class PaperToExcelPipeline:
    def __init__(self):

        self.wb = Workbook()
        del self.wb['Sheet']
        self.default_style = config_style()
        self.ws_dict = {p['title']: self.wb.create_sheet(title=p['title']) for p in paper_list}
        self.save_path = 'papers.xlsx'

    def process_item(self, item: Dict, spider):
        title = item['title']
        ws = self.ws_dict[title]
        paper = item['paper']

        ws.append([paper])
        return item

    def open_spider(self, spider: scrapy.Spider):
        spider.log('spider open')
        pass

    def apply_style(self):
        for ws in self.ws_dict.values():
            for r in ws.rows:
                c: Cell
                for c in r:
                    c.style = self.default_style

            c: Tuple[Cell]
            for c in ws.columns:
                column = c[0].column
                max_len = 0
                r: Cell
                for r in c:
                    if r.value is None:
                        break
                    max_len = max(max_len, len(r.value))
                adjusted_width = max_len * 1.1
                print('adjusted_width', adjusted_width)
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    def close_spider(self, spider):
        spider.log('spider close')
        self.apply_style()
        self.wb.save(filename=self.save_path)
        self.wb.close()
